# Tabel incarcari verticale
from openpyxl import Workbook
wb = Workbook() # current workbook
ws = wb.active  # default worksheet

#==========================================#==========================================
# INITIALISE STAGE
# Initialize all the values for table header, formatting, etc
#==========================================#==========================================
measure_units = {'m1': '[m]', \
                 'm2': '[m^2]', \
                 'm3': '[daN]'}
th_text = {'A1': 'ÎNCĂRCĂRI VERTICALE PE ȘPALEȚI', \
           'A2':'Diafrag',\
           'B2':'Șpalet',\
           'C2':'Grosime P', 'C3': measure_units['m1'],\
           'D2':'Aaf peste parter', 'D3': measure_units['m2'],\
           'E2':'Grosime E', 'E3': measure_units['m1'],\
           'F2':'Aaf peste etaj', 'F3': measure_units['m2'],\
           'G2':'L. Șpaleti (parter)', 'G3': measure_units['m1'],\
           'H2':'H. Pereti (parter)', 'H3': measure_units['m1'],\
           'I2':'L. Parapeti (parter)', 'I3': measure_units['m1'],\
           'J2':'H. Parapeti (parter)', 'J3': measure_units['m1'],\
           'K2':'L. Centuri (parter)', 'K3': measure_units['m1'],\
           'L2':'H. Centuri (parter)', 'L3': measure_units['m1'],\
           'M2':'L. Șpaleti (etaj)', 'M3': measure_units['m1'],\
           'N2':'H. Pereti (etaj)', 'N3': measure_units['m1'],\
           'O2':'L. Parapeti (etaj', 'O3': measure_units['m1'],\
           'P2':'H. Parapeti (etaj)', 'P3': measure_units['m1'],\
           'Q2':'L. Centuri (etaj)', 'Q3': measure_units['m1'],\
           'R2':'H. Centuri (etaj)', 'R3': measure_units['m1'],\
           'S2':'L. Atic', 'S3': measure_units['m1'],\
           'T2':'H. Atic', 'T3': measure_units['m1'],\
           'U2':'L. Centuri (atic)', 'U3': measure_units['m1'],\
           'V2':'H. Centuri (atic)', 'V3': measure_units['m1'],\
           'W2':'Ned', 'W3': measure_units['m3'],\
           'X2':'Nsd', 'X3': measure_units['m3'],\
           }

#==========================================#==========================================
# STAGE ONE
# CREATE THE TABLE
#==========================================#==========================================
#A1:X1 (CAP TABEL)
ws.merge_cells("A1:X1")
#A2:A3 diafrag
ws.merge_cells("A2:A3")
#B2:B3 SPALETI
ws.merge_cells("B2:B3")
#==========================================#==========================================
# STAGE TWO
# FORMAT THE TABLE
#==========================================#==========================================

#==========================================#==========================================
# STAGE <TBD>
# Fill in the values
#ex. how to assing a value to a cell
# ws['A1'] = th_text['A1']
#==========================================#==========================================
# TABLE HEADER
#-----------------#-----------------#-----------------
for i in th_text:
    ws[i]= th_text[i]
#-----------------#-----------------#-----------------


#==========================================#==========================================
# STAGE <TBD>
# Save the Worksheets.
#==========================================#==========================================
wb.save("ivps.xlsx")